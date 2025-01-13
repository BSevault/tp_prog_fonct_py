# lecture et écriture de fichiers Excel
# https://openpyxl.readthedocs.io/en/stable/ 
# https://openpyxl.readthedocs.io/en/stable/tutorial.html#create-a-workbook 

# nécessite :
# pip install openpyxl 
# pip install pillow   si usage d'image dans le document excel.

from openpyxl import Workbook

# Créer un document Excel
wb = Workbook() 
ws = wb.active  #récupèration de la feuille 0 qui existe par défaut
# ws1 = wb.create_sheet("Mysheet") # insert at the end (default)  
d = ws.cell(row=4, column=2, value=10) # écrire une valeur
ws['A1'] = 'hello world' #autre façon d'écrire une valeur

wb.save("essaiOpenpyxl.xlsx")
wb.close()


# Lire un document Excel
wb1 = Workbook('essaiOpenpyxl.xlsx')
# wb1 = load_workbook('essaiOpenpyxl.xlsx')
ws1 = wb1.active
st = ws['A1']
print(st.value) # st.value est le contenu de la cellule
wb1.close()


